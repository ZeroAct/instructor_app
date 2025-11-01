"""File parsing utilities with PaddleOCR support."""

import io
import json
import os
from pathlib import Path
from typing import Optional

# Check if file upload feature is enabled
def is_file_upload_enabled() -> bool:
    """Check if file upload feature is enabled via environment variable."""
    return os.getenv("ENABLE_FILE_UPLOAD", "true").lower() in ("true", "1", "yes")


class FileParser:
    """File parser with OCR and text extraction capabilities."""

    def __init__(self, config_path: str = "config.json"):
        """Initialize file parser with configuration."""
        self.config = self._load_config(config_path)
        self._ocr = None
        self._pdf_available = False
        self._docx_available = False
        self._xlsx_available = False
        
        # Cache imports to avoid repeated overhead
        self._PyPDF2 = None
        self._docx = None
        self._openpyxl = None
        
        # Check available libraries
        self._check_dependencies()

    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        # Try to find config in multiple locations
        possible_paths = [
            config_path,
            os.path.join(os.getcwd(), config_path),
            os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", config_path),
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                with open(path, "r") as f:
                    return json.load(f)
        
        # Return default config if file not found
        return {
            "file_upload": {
                "enabled": True,
                "max_file_size_mb": 10,
                "allowed_extensions": [
                    ".txt", ".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".png", 
                    ".bmp", ".gif", ".tiff", ".csv", ".xlsx", ".xls", ".json", 
                    ".xml", ".html", ".md", ".rtf"
                ]
            },
            "ocr": {
                "backend": "paddleocr",
                "paddleocr": {
                    "use_angle_cls": True,
                    "lang": "en",
                    "use_gpu": False,
                    "show_log": False
                },
                "fallback_to_text": True,
                "image_extensions": [".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp"]
            },
            "text_extraction": {
                "encoding": "utf-8",
                "fallback_encodings": ["latin-1", "cp1252", "iso-8859-1"]
            },
            "pdf_processing": {
                "extract_images": True,
                "ocr_images": True
            }
        }

    def _check_dependencies(self):
        """Check which optional dependencies are available."""
        try:
            import PyPDF2
            self._PyPDF2 = PyPDF2
            self._pdf_available = True
        except ImportError:
            pass
        
        try:
            import docx
            self._docx = docx
            self._docx_available = True
        except ImportError:
            pass
        
        try:
            import openpyxl
            self._openpyxl = openpyxl
            self._xlsx_available = True
        except ImportError:
            pass

    def _get_ocr(self):
        """Lazy load PaddleOCR."""
        if self._ocr is None:
            try:
                from paddleocr import PaddleOCR
                import inspect
                
                ocr_config = self.config.get("ocr", {}).get("paddleocr", {})
                
                # Get valid parameters for PaddleOCR.__init__
                sig = inspect.signature(PaddleOCR.__init__)
                valid_params = set(sig.parameters.keys()) - {'self'}
                
                # Build kwargs with only valid parameters
                kwargs = {}
                
                # Common parameters with their config keys and defaults
                param_mapping = {
                    'use_angle_cls': ('use_angle_cls', True),
                    'lang': ('lang', 'en'),
                    'use_gpu': ('use_gpu', False),  # May not be valid in newer versions
                    'show_log': ('show_log', False),
                    'det_model_dir': ('det_model_dir', None),
                    'rec_model_dir': ('rec_model_dir', None),
                    'cls_model_dir': ('cls_model_dir', None),
                }
                
                for param_name, (config_key, default_value) in param_mapping.items():
                    if param_name in valid_params:
                        value = ocr_config.get(config_key, default_value)
                        # Only add non-None values
                        if value is not None:
                            kwargs[param_name] = value
                
                self._ocr = PaddleOCR(**kwargs)
            except ImportError:
                raise ImportError(
                    "PaddleOCR is not installed. Install it with: pip install paddleocr"
                )
        return self._ocr

    def parse_file(self, file_content: bytes, filename: str) -> str:
        """
        Parse file content and extract text.
        
        Args:
            file_content: Binary content of the file
            filename: Name of the file (used to determine type)
            
        Returns:
            Extracted text content
        """
        if not is_file_upload_enabled():
            raise ValueError("File upload feature is disabled")
        
        ext = Path(filename).suffix.lower()
        
        # Check if extension is allowed
        allowed_extensions = self.config.get("file_upload", {}).get("allowed_extensions", [])
        if allowed_extensions and ext not in allowed_extensions:
            raise ValueError(f"File extension {ext} is not allowed")
        
        # Check file size
        max_size_mb = self.config.get("file_upload", {}).get("max_file_size_mb", 10)
        if len(file_content) > max_size_mb * 1024 * 1024:
            raise ValueError(f"File size exceeds maximum of {max_size_mb}MB")
        
        # Route to appropriate parser
        if ext in self.config.get("ocr", {}).get("image_extensions", []):
            return self._parse_image(file_content)
        elif ext == ".pdf":
            return self._parse_pdf(file_content)
        elif ext in [".doc", ".docx"]:
            return self._parse_docx(file_content)
        elif ext in [".xls", ".xlsx"]:
            return self._parse_xlsx(file_content)
        elif ext in [".csv"]:
            return self._parse_csv(file_content)
        elif ext in [".json"]:
            return self._parse_json(file_content)
        elif ext in [".xml", ".html"]:
            return self._parse_markup(file_content)
        else:
            # Fallback to text extraction
            return self._parse_text(file_content)

    def _parse_image(self, content: bytes) -> str:
        """Parse image file using OCR."""
        import tempfile
        
        try:
            ocr = self._get_ocr()
            
            # Save to temporary file for PaddleOCR using context manager
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            
            try:
                result = ocr.ocr(tmp_path, cls=True)
                
                # Extract text from OCR result
                text_lines = []
                if result and result[0]:
                    for line in result[0]:
                        if line and len(line) > 1:
                            text_lines.append(line[1][0])
                
                return "\n".join(text_lines) if text_lines else ""
            finally:
                # Clean up temp file - guaranteed to execute
                try:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                except Exception:
                    pass  # Ignore cleanup errors
                    
        except Exception as e:
            if self.config.get("ocr", {}).get("fallback_to_text", True):
                # Try to extract any text metadata
                return f"[Image file - OCR failed: {str(e)}]"
            raise

    def _parse_pdf(self, content: bytes) -> str:
        """Parse PDF file."""
        if not self._pdf_available:
            # Don't fallback to text parsing for binary PDF files
            return "[PDF file detected but PyPDF2 is not installed. Install it with: pip install PyPDF2]"
        
        try:
            pdf_reader = self._PyPDF2.PdfReader(io.BytesIO(content))
            text_parts = []
            
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text and text.strip():
                    text_parts.append(text)
            
            extracted_text = "\n\n".join(text_parts)
            
            # If PDF has no text and OCR is enabled, try OCR on images
            if not extracted_text.strip() and self.config.get("pdf_processing", {}).get("ocr_images", True):
                try:
                    # Try to extract images and OCR them
                    return "[PDF with no extractable text - OCR on images not yet implemented]"
                except Exception:
                    pass
            
            return extracted_text if extracted_text.strip() else "[Empty PDF]"
            
        except Exception as e:
            # Return error message instead of fallback to text parsing
            return f"[PDF parsing failed: {str(e)}. The file may be corrupted or password-protected.]"

    def _parse_docx(self, content: bytes) -> str:
        """Parse DOCX file."""
        if not self._docx_available:
            # Don't fallback to text parsing for binary DOCX files
            return "[DOCX file detected but python-docx is not installed. Install it with: pip install python-docx]"
        
        try:
            doc = self._docx.Document(io.BytesIO(content))
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else "[Empty document]"
            
        except Exception as e:
            # Return error message instead of fallback to text parsing
            return f"[DOCX parsing failed: {str(e)}. The file may be corrupted or in an unsupported format.]"

    def _parse_xlsx(self, content: bytes) -> str:
        """Parse Excel file."""
        if not self._xlsx_available:
            # Don't fallback to text parsing for binary Excel files
            return "[Excel file detected but openpyxl is not installed. Install it with: pip install openpyxl]"
        
        try:
            workbook = self._openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
                
                text_parts.append("")  # Empty line between sheets
            
            return "\n".join(text_parts) if text_parts else "[Empty spreadsheet]"
            
        except Exception as e:
            # Return error message instead of fallback to text parsing
            return f"[Excel parsing failed: {str(e)}. The file may be corrupted or in an unsupported format.]"

    def _parse_csv(self, content: bytes) -> str:
        """Parse CSV file."""
        try:
            import csv
            
            # Try to decode with configured encoding
            text = self._decode_bytes(content)
            
            # Parse CSV
            reader = csv.reader(io.StringIO(text))
            rows = []
            for row in reader:
                rows.append("\t".join(row))
            
            return "\n".join(rows) if rows else "[Empty CSV]"
            
        except Exception as e:
            if self.config.get("ocr", {}).get("fallback_to_text", True):
                return self._parse_text(content)
            raise

    def _parse_json(self, content: bytes) -> str:
        """Parse JSON file."""
        try:
            text = self._decode_bytes(content)
            data = json.loads(text)
            
            # Pretty print JSON
            return json.dumps(data, indent=2)
            
        except Exception as e:
            if self.config.get("ocr", {}).get("fallback_to_text", True):
                return self._parse_text(content)
            raise

    def _parse_markup(self, content: bytes) -> str:
        """Parse XML/HTML file."""
        try:
            text = self._decode_bytes(content)
            
            # Try to extract text from HTML/XML tags
            try:
                from html.parser import HTMLParser
                
                class TextExtractor(HTMLParser):
                    def __init__(self):
                        super().__init__()
                        self.text_parts = []
                    
                    def handle_data(self, data):
                        if data.strip():
                            self.text_parts.append(data.strip())
                
                parser = TextExtractor()
                parser.feed(text)
                
                return "\n".join(parser.text_parts) if parser.text_parts else text
            except Exception:
                return text
                
        except Exception as e:
            if self.config.get("ocr", {}).get("fallback_to_text", True):
                return self._parse_text(content)
            raise

    def _parse_text(self, content: bytes) -> str:
        """Parse plain text file with encoding detection."""
        return self._decode_bytes(content)

    def _decode_bytes(self, content: bytes) -> str:
        """Decode bytes to string with fallback encodings."""
        encodings = [self.config.get("text_extraction", {}).get("encoding", "utf-8")]
        encodings.extend(
            self.config.get("text_extraction", {}).get("fallback_encodings", [])
        )
        
        for encoding in encodings:
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                continue
        
        # Last resort: decode with errors='replace'
        return content.decode("utf-8", errors="replace")


# Singleton instance
_parser_instance: Optional[FileParser] = None


def get_file_parser() -> FileParser:
    """Get or create file parser singleton."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = FileParser()
    return _parser_instance
